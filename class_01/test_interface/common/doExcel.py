# -*- coding : utf-8 -*-
# @Time      :2019/3/27 22:49
# @Author    : py15期   lemon_huihui
# @File      : doExcel.py
from openpyxl import load_workbook

from class_01.test_interface.common import http_request


class Case():
    '''
    封装一个case类,把所有字段放在对象属性里， 初始化所有字段 作为属性，
    方便存取 用例的 value。后面实例一个对象就是一条用例。

    '''
    def __init__(self):
        self.case_id = None
        self.title = None
        self.url = None
        self.data = None
        self.method = None
        self.expected = None
        self.actual = None
        self.result = None
        self.check_sql = None

class DoExcel():
    def __init__(self,filename,sheet_name):
        self.filename = filename
        self.sheet_name = sheet_name
        self.wb = load_workbook(self.filename)####打开excel文件
        self.sheet = self.wb[self.sheet_name]

    def get_cases(self):
        max_row = self.sheet.max_row

        cases=[]#把所有对象 最后又存放在 一个list 里
        for row in range(2,max_row+1):
            case=Case()##必须实例一个新对象

            case.case_id = self.sheet.cell(row=row,column=1).value #读取出所有对应value 附给 对象的属性
            case.title = self.sheet.cell(row=row,column=2).value
            case.url = self.sheet.cell(row=row,column=3).value
            case.data = self.sheet.cell(row=row,column=4).value
            case.method = self.sheet.cell(row=row,column=5).value
            case.expected = self.sheet.cell(row=row,column=6).value
            case.check_sql= self.sheet.cell(row=row,column=9).value

            cases.append(case)##最后把一个对象 放进一个list
        self.wb.close()
        return cases
    def get_dict_cases(self):
        max_row = self.sheet.max_row

        dict_cases = []  # 把所有对象 最后又存放在 一个list 里
        for row in range(2, max_row + 1):
            dict = {}#定义一个空字典  存放每行 每条用例的 value

            dict['case_id'] = self.sheet.cell(row=row,column=1).value #读取出所有对应value 附给 对象的属性
            dict['title'] = self.sheet.cell(row=row,column=2).value
            dict['url'] = self.sheet.cell(row=row,column=3).value
            dict['data'] = self.sheet.cell(row=row,column=4).value
            dict['method'] = self.sheet.cell(row=row,column=5).value
            dict['expected'] = self.sheet.cell(row=row,column=6).value
            dict['check_sql'] = self.sheet.cell(row=row, column=9).value
            dict_cases.append(dict)
        self.wb.close()
        return dict_cases

    def write_result(self,row,actual,result):#写入结果到
        self.sheet.cell(row,7).value=actual
        self.sheet.cell(row,8).value=result
        self.wb.save(self.filename)
        self.wb.close()

if __name__ == '__main__':
    # from class_01.test_interface import contants
    do_excel=DoExcel(filename='cases.xlsx',sheet_name='login')
    cases=do_excel.get_cases()
    http2= http_request.HttpRequest2()
    for case in cases:
        print(case.__dict__)
        print(case.case_id,case.title)
        params = {"mobilephone": "15810447878", "pwd": "123456"}
        resp=http2.request(method=case.method,url=case.url,data=eval(case.data))
        print('响应结果是：',resp.text)
        if resp.text == case.expected:
            do_excel.write_result(row=case.case_id+1,actual=resp.text,result='PASS')
        else:
            do_excel.write_result(row=case.case_id+1,actual=resp.text,result='FAIL')
        #充值
    params = {"mobilephone": "15810447878", "amount": "1000"}
    resp2 = http2.request('POST','http://test.lemonban.com/futureloan/mvc/api/member/recharge', data=params)
    http2.close()
    print(resp2.status_code)
    print(resp2.text)
    print(resp2.request._cookies)









